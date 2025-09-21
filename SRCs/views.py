import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from .forms import formularioUnidade, formularioUser, formularioEnvio, UploadFaturaForm, FormularioEditarUsuario, formularioItemEnvio
from .models import Unidade, Usuario, Envio, Rateio, ItemEnvio
from django.contrib import messages
import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from urllib.parse import urlencode
from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum
import json
from django.http import HttpResponse
from django.db.models import Q
from django.forms import inlineformset_factory
import pandas as pd
from .utils import gerar_pdf_declaracao  # Corrigir nome e acentuação
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db.models import Sum



# View da página inicial
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        senha = request.POST.get('password')

        user = authenticate(request, username=username, password=senha)

        if user is not None:
            login(request, user)
            return redirect('home')  # ou a página inicial do sistema
        else:
            messages.error(request, "Usuário ou senha inválidos.")

    return render(request, 'SRCs/login.html')

# View da home após login
@login_required
def home(request):
    usuario_logado = Usuario.objects.get(user=request.user)

    ultimos_envios = Envio.objects.all().order_by('-data_solicitacao')[:4]
    pendentes = Envio.objects.filter(usuario_destinatario=usuario_logado, status='Pendente')

    return render(request, 'SRCs/home.html', {
        'ultimos_envios': ultimos_envios,
        'pendentes': pendentes,
    })

# View que lista todos os usuários (acesso apenas para quem tem permissão)
@login_required
@permission_required('SRCs.editar_usuario', raise_exception=True)
def user(request):
    # Se for POST, estamos editando um usuário via modal
    if request.method == "POST":
        usuario_id = request.POST.get('usuario_id')
        usuario = get_object_or_404(Usuario, pk=usuario_id)

        # Atualiza os campos básicos do usuário
        usuario.user.first_name = request.POST.get('nome')
        usuario.user.email = request.POST.get('email')

        # Atualiza unidade se enviada
        unidade_id = request.POST.get('unidade')
        if unidade_id:
            try:
                usuario.unidade = Unidade.objects.get(id=unidade_id)
            except Unidade.DoesNotExist:
                messages.error(request, "Unidade selecionada não existe.")
        else:
            usuario.unidade = None

        usuario.user.save()
        usuario.save()
        messages.success(request, "Usuário atualizado com sucesso!")
        return redirect('user')  # volta para a mesma página

    # GET: lista os usuários
    usuarios = Usuario.objects.filter(ativo=True).select_related('user', 'unidade')
    unidades = Unidade.objects.all()  # 🔑 manda todas as unidades para o template

    return render(request, 'SRCs/user.html', {
        'usuarios': usuarios,
        'unidades': unidades
    })

# View para cadastrar novos usuários
@login_required
@permission_required('SRCs.cadastrar_usuario', raise_exception=True)
def cadastro_user(request):
    if request.method == 'POST':
        form = formularioUser(request.POST, request.FILES)  
        if form.is_valid():
            usuario = form.save(commit=False)   # cria o objeto mas não salva ainda
            usuario.save()                      # salva o Usuario no banco
            usuario.atribuir_grupo()            # chama seu método para garantir o Group
            return redirect('cadastro_user') 
    else:
        form = formularioUser()
    
    return render(request, 'SRCs/form_user.html', {'form': form})

from django.contrib import messages

def importar_unidade(request):
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        arquivo = request.FILES['arquivo_excel']
        cnpjs_duplicados = []
        unidades_importadas = 0  # contador de novas unidades
        try:
            # Ler o Excel
            df = pd.read_excel(arquivo)

            # Campos esperados no Excel
            campos_esperados = [
                'cnpj', 'centro_custo', 'cep', 'bairro', 'rua', 'numero',
                'shopping', 'cidade', 'estado', 'regional', 'numero_unidade', 'empresa'
            ]

            # Verificar se todos os campos estão presentes
            for campo in campos_esperados:
                if campo not in df.columns:
                    messages.error(request, f"Coluna '{campo}' não encontrada no Excel.")
                    return redirect('cadastro_unidade')

            # Criar as unidades
            for _, row in df.iterrows():
                cnpj = str(row['cnpj']).strip()
                if Unidade.objects.filter(cnpj=cnpj).exists():
                    cnpjs_duplicados.append(cnpj)
                    continue  # pula duplicados
                Unidade.objects.create(
                    cnpj=cnpj,
                    centro_custo=row['centro_custo'],
                    cep=row['cep'],
                    bairro=row['bairro'],
                    rua=row['rua'],
                    numero=row['numero'],
                    shopping=row['shopping'],
                    cidade=row['cidade'],
                    estado=row['estado'],
                    regional=row['regional'],
                    numero_unidade=row['numero_unidade'],
                    empresa=row['empresa']
                )
                unidades_importadas += 1

            # Mensagens condicionais
            if unidades_importadas > 0:
                messages.success(request, f"{unidades_importadas} unidades importadas com sucesso!")

            if cnpjs_duplicados:
                mensagens = ", ".join(cnpjs_duplicados)
                messages.warning(request, f"As unidades com CNPJs {mensagens} já existem e não foram importadas.")

        except Exception as e:
            messages.error(request, f"Erro ao importar: {e}")

    return redirect('cadastro_unidade')

# View para cadastrar unidades
@login_required
@permission_required('SRCs.cadastrar_unidade', raise_exception=True)
def cadastro_unidade(request):
    id_unidade = request.GET.get('ids') or request.POST.get('id')  
    unidade = Unidade.objects.filter(id=id_unidade).first() if id_unidade else None

    if request.method == 'POST':
        form = formularioUnidade(request.POST, instance=unidade)

        # Verificação de CNPJ duplicado
        cnpj = request.POST.get('cnpj')
        if cnpj:
            # Se estamos editando, ignoramos o próprio registro
            duplicado = Unidade.objects.filter(cnpj=cnpj)
            if unidade:
                duplicado = duplicado.exclude(id=unidade.id)

            if duplicado.exists():
                messages.error(request, f"O CNPJ {cnpj} já está em uso.")
                modo = 'editar' if unidade else 'cadastrar'
                return render(request, 'SRCs/form_und.html', {'form': form, 'modo': modo})

        if form.is_valid():
            form.save()
            messages.success(request, "Unidade cadastrada com sucesso!")
            return redirect('cadastro_unidade')
    else:
        form = formularioUnidade(instance=unidade)

    modo = 'editar' if unidade else 'cadastrar'
    return render(request, 'SRCs/form_und.html', {'form': form, 'modo': modo})

# View que lista todas as unidades
@login_required
@permission_required('SRCs.editar_unidade', raise_exception=True)
def unidade(request):
    unidades = Unidade.objects.filter(excluida=False)
    return render(request, 'SRCs/unidade.html', {'unidades': unidades})

ItemEnvioFormSet = inlineformset_factory(
    Envio, ItemEnvio,
    form=formularioItemEnvio,
    extra=1,  # quantidade inicial de formulários de item
    can_delete=True
)

# View para cadastrar um envio (ex: envio pelos Correios)
@login_required
@permission_required('SRCs.cadastrar_envio', raise_exception=True)
def cadastro_envio(request):
    if request.method == 'POST':
        form = formularioEnvio(request.POST)
        formset = ItemEnvioFormSet(request.POST, prefix='form')

        if form.is_valid() and formset.is_valid():
            envio = form.save(commit=False)
            envio.user = request.user
            envio.save()

            itens = formset.save(commit=False)
            for item in itens:
                item.envio = envio
                item.save()

            if request.POST.get("gerar_pdf") == "sim":
                return gerar_pdf_declaracao(envio)

            return redirect('home')
    else:
        form = formularioEnvio()
        formset = ItemEnvioFormSet(prefix='form')

    return render(request, 'SRCs/form_envio.html', {
        'form': form,
        'formset': formset
    })


# Função auxiliar para converter valores em decimal de forma segura
def safe_decimal(valor):
    try:
        if valor is None:
            return Decimal('0.00')

        valor_str = str(valor).strip()

        # Trata valores vazios ou com traço
        if valor_str in ('', '-', '–'):
            return Decimal('0.00')

        # Formato brasileiro: converte R$ e vírgula para ponto
        valor_str = valor_str.replace('R$', '').replace('.', '').replace(',', '.')
        return Decimal(valor_str)
    
    except (InvalidOperation, ValueError) as e:
        raise ValueError(f"Valor inválido para Decimal: {valor}")

# View que faz o processamento da fatura (rateio)
@login_required
@permission_required('SRCs.consultar_rateio', raise_exception=True)
def rateio(request):
    if request.method == 'POST':
        form = UploadFaturaForm(request.POST, request.FILES)
        if form.is_valid():
            fatura_file = request.FILES['fatura']
            nome_arquivo = fatura_file.name.rsplit('.', 1)[0]
            try:
                # Abre o arquivo Excel enviado
                wb = openpyxl.load_workbook(fatura_file, data_only=True)
                ws = wb.active

                erros = []       # Lista de erros encontrados
                importados = 0   # Contador de registros importados

                # Começa a ler da linha 6 em diante
                for row in ws.iter_rows(min_row=6, values_only=True):
                    print(f"Lendo linha: {row}")

                    if row[0] is None or not isinstance(row[0], (int, float)):
                        print(f"Encerrando leitura por linha inválida: {row[0]}")
                        break

                    etiqueta_valor = str(row[8]).strip().upper()
                    print(f"Etiqueta buscada (ajustada): '{etiqueta_valor}'")

                    envio = Envio.objects.filter(etiqueta__iexact=etiqueta_valor).first()
                    print(f"Envio encontrado? {'Sim' if envio else 'Não'}")

                    if not envio:
                        print(f"[INFO] Etiqueta '{etiqueta_valor}' não encontrada. Criando rateio sem envio.")



                    # Aqui vai o try de criação do rateio
                    try:

                        Rateio.objects.filter(etiqueta_original=etiqueta_valor).delete()


                        Rateio.objects.create(
                            etiqueta=envio,
                            etiqueta_original=etiqueta_valor,
                            fatura=nome_arquivo,
                            titular_cartao=row[1],
                            servico=row[3],
                            data_postagem = (
                                datetime.datetime.strptime(row[4], '%d/%m/%Y').date()
                                if isinstance(row[4], str) and row[4].strip()
                                else row[4] if isinstance(row[4], datetime.datetime)
                                else None
                            ),
                            servico_adicionais=safe_decimal(row[5]),
                            unidade_postagem=row[6],
                            valor_declarado=safe_decimal(row[15]),
                            valor_unitario=safe_decimal(row[11]),
                            peso=safe_decimal(row[10]),
                            desconto=safe_decimal(row[12]),
                            valor_liquido=safe_decimal(row[14]),
                        )
                        print(f"Rateio criado com sucesso para a etiqueta {etiqueta_valor}")
                        importados += 1
                    except Exception as e:
                        erros.append(f"Erro ao importar etiqueta '{etiqueta_valor}': {str(e)}")

                # Mensagem de sucesso/erro
                if importados:
                    messages.success(request, f'{importados} registros importados com sucesso.')
                if erros:
                    for erro in erros:
                        messages.warning(request, erro)
                
                return redirect('rateio')

            except Exception as e:
                messages.error(request,  f'Erro ao processar o arquivo: {str(e)}')
    else:
        form = UploadFaturaForm()
    
    # Busca envios e rateios para exibição na página
    envios = Envio.objects.select_related('user', 'remetente', 'destinatario').all().order_by('-data_solicitacao')
    rateios = Rateio.objects.select_related('etiqueta').all()

    # Para evitar repetição de etiquetas
    etiquetas_processadas = set()

    # Associa etiquetas não vinculadas no banco
    for rateio in rateios:
        if not rateio.etiqueta:
            envio = Envio.objects.filter(etiqueta=rateio.etiqueta_original).first()
            if envio:
                rateio.etiqueta = envio
                rateio.save()

    dados_completos = []

    for envio in envios:
        rateio = Rateio.objects.filter(etiqueta=envio).order_by('-id').first()

        # Conteúdo concatenado de todos os itens do envio
        conteudo_envio = ", ".join([item.conteudo for item in envio.itens.all()])

        # Cartão de postagem via Usuario
        cartao_postagem = getattr(getattr(envio.user, 'usuario', None), 'cartao_postagem', '')

        dados_completos.append({
            'fatura': rateio.fatura if rateio else 'PENDENTE',
            'solicitante': envio.user.get_full_name() or envio.user.username,
            'motivo': getattr(envio, 'motivo', ''),
            'conteudo': conteudo_envio,
            'cartao_postagem': cartao_postagem,
            'titular_cartao': getattr(rateio, 'titular_cartao', ''),
            'servico': getattr(rateio, 'servico', ''),
            'numero_autorizacao': getattr(envio, 'numero_autorizacao', ''),
            'etiqueta': getattr(envio, 'etiqueta', ''),
            'data_postagem': getattr(rateio, 'data_postagem', ''),
            'unidade_postagem': getattr(rateio, 'unidade_postagem', ''),
            'remetente': getattr(envio.remetente, 'shopping', ''),
            'destinatario': getattr(envio.destinatario, 'shopping', ''),
            'valor_declarado': getattr(rateio, 'valor_declarado', ''),
            'valor_unitario': getattr(rateio, 'valor_unitario', ''),
            'quantidade': getattr(envio, 'quantidade', ''),  # se tiver no envio ou somar itens
            'peso': getattr(rateio, 'peso', ''),
            'servico_adicionais': getattr(rateio, 'servico_adicionais', ''),
            'valor_liquido': getattr(rateio, 'valor_liquido', ''),
            'desconto': getattr(rateio, 'desconto', ''),
            'centro_custo': getattr(envio.destinatario, 'centro_custo', ''),
            'empresa': getattr(envio.destinatario, 'empresa', ''),
        })
        etiquetas_processadas.add(envio.etiqueta)

    # Adiciona os rateios que não têm envio correspondente
    for rateio in rateios:
        if rateio.etiqueta and rateio.etiqueta.etiqueta in etiquetas_processadas:
            continue 
        
        dados_completos.append({
            'fatura': rateio.fatura,
            'solicitante': '',
            'motivo': '',
            'cartao_postagem': '',
            'titular_cartao': rateio.titular_cartao,
            'servico': rateio.servico,
            'numero_autorizacao': '',
            'etiqueta': rateio.etiqueta.etiqueta if rateio.etiqueta else rateio.etiqueta_original,
            'data_postagem': rateio.data_postagem,
            'unidade_postagem': rateio.unidade_postagem,
            'remetente': '',
            'destinatario': '',
            'valor_declarado': rateio.valor_declarado,
            'valor_unitario': rateio.valor_unitario,
            'quantidade': '',
            'peso': rateio.peso,
            'servico_adicionais': rateio.servico_adicionais,
            'valor_liquido': rateio.valor_liquido,
            'desconto': rateio.desconto,
            'centro_custo': '',
            'empresa': '',
        })

    # Renderiza a página com os dados e o formulário de upload
    return render(request, 'SRCs/rateio.html', {'form': form, 'dados': dados_completos}) 

def exportar_rateio(request):
    #Mostra os dados completos 
    envios = Envio.objects.select_related('user', 'remetente', 'destinatario').all().order_by('-data_solicitacao')
    rateios = Rateio.objects.select_related('etiqueta').all()

    etiquetas_processadas = set()
    dados_completos = []

    for envio in envios:
        rateio = Rateio.objects.filter(etiqueta=envio).order_by('-id').first()
        dados_completos.append({
            'fatura': rateio.fatura if rateio else 'PENDENTE',
            'solicitante': envio.user.get_full_name() or envio.user.username,
            'motivo': envio.motivo,  # <- agora é o campo correto do Envio
            'itens': ", ".join([f"{item.conteudo} ({item.quantidade}x)" for item in envio.itens.all()]), # lista de itens
            'cartao_postagem': getattr(envio.user, 'cartao_postagem', ''),
            'titular_cartao': rateio.titular_cartao if rateio else '',
            'servico': rateio.servico if rateio else '',
            'numero_autorizacao': envio.numero_autorizacao,
            'etiqueta': envio.etiqueta,
            'data_postagem': rateio.data_postagem if rateio else '',
            'unidade_postagem': rateio.unidade_postagem if rateio else '',
            'remetente': envio.remetente.shopping,
            'destinatario': envio.destinatario.shopping,
            'valor_declarado': rateio.valor_declarado if rateio else '',
            'valor_unitario': rateio.valor_unitario if rateio else '',
            'quantidade': envio.itens.aggregate(total=Sum('quantidade'))['total'] or '',
            'peso': rateio.peso if rateio else '',
            'servico_adicionais': rateio.servico_adicionais if rateio else '',
            'valor_liquido': rateio.valor_liquido if rateio else '',
            'desconto': rateio.desconto if rateio else '',
            'centro_custo': envio.destinatario.centro_custo,
            'empresa': envio.destinatario.empresa,
        })

    for rateio in rateios:
        if rateio.etiqueta and rateio.etiqueta.etiqueta in etiquetas_processadas:
            continue
        dados_completos.append({
            'fatura': rateio.fatura,
            'solicitante': '',
            'motivo': '',
            'cartao_postagem': '',
            'titular_cartao': rateio.titular_cartao,
            'servico': rateio.servico,
            'numero_autorizacao': '',
            'etiqueta': rateio.etiqueta.etiqueta if rateio.etiqueta else rateio.etiqueta_original,
            'data_postagem': rateio.data_postagem,
            'unidade_postagem': rateio.unidade_postagem,
            'remetente': '',
            'destinatario': '',
            'valor_declarado': rateio.valor_declarado,
            'valor_unitario': rateio.valor_unitario,
            'quantidade': '',
            'peso': rateio.peso,
            'servico_adicionais': rateio.servico_adicionais,
            'valor_liquido': rateio.valor_liquido,
            'desconto': rateio.desconto,
            'centro_custo': '',
            'empresa': '',
        })
    
    #Cria o Excal
    wb = Workbook() 
    ws = wb.active
    ws.title = "Rateio"

    #cabeçalho
    colunas = list(dados_completos[0].keys()) if dados_completos else [] 
    ws.append(colunas)

    #linhas 
    for item in dados_completos:
        ws.append([item.get(c, '') for c in colunas])

    #retorna com dowload 
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rateio_completo.xlsx'
    wb.save(response)
    return response

# View do painel de gráficos
@login_required
@permission_required('SRCs.consultar_dashboard', raise_exception=True)
def dashboard(request):

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    unidade_id = request.GET.get('unidade')
    usuario_id = request.GET.get('usuario')

    rateios = Rateio.objects.exclude(data_postagem=None)
    envios = Envio.objects.all()

    # Aplicar filtros de data
    if data_inicio:
        rateios = rateios.filter(data_postagem__gte=data_inicio)
        envios = envios.filter(data_solicitacao__gte=data_inicio)
    if data_fim:
        rateios = rateios.filter(data_postagem__lte=data_fim)
        envios = envios.filter(data_solicitacao__lte=data_fim)
    if unidade_id:
        envios = envios.filter(remetente_id=unidade_id)  # ou destinatario_id
    if usuario_id:
        envios = envios.filter(user_id=usuario_id)

    # ===== Gráfico 1: Quantidade de envios por mês =====
    dados = (
        rateios
        .annotate(mes=TruncMonth('data_postagem'))
        .values('mes')
        .annotate(qtd=Count('id'))
        .order_by('mes')
    )
    labels = [f"{item['mes'].strftime('%b/%Y')}" for item in dados]
    valores = [item['qtd'] for item in dados]

    context = {
        'labels': json.dumps(labels),
        'valores': json.dumps(valores),
    }

    # ===== Gráfico 2: Envios por unidade remetente =====
    envios_por_remetente = (
        envios
        .values('remetente__shopping')
        .annotate(total=Count('etiqueta'))
        .order_by('-total')
    )
    remetentes = [item['remetente__shopping'] or "Não informado" for item in envios_por_remetente]
    envios_qtd = [item['total'] for item in envios_por_remetente]
    context['remetentes'] = json.dumps(remetentes)
    context['qtd_envios'] = json.dumps(envios_qtd)

    # ===== Gráfico 3: Envios por unidade destinatária =====
    envios_por_destinatario = (
        envios
        .values('destinatario__shopping')
        .annotate(total=Count('etiqueta'))
        .order_by('-total')
    )
    destinatarios = [item.get('destinatario__shopping') or 'Não informado' for item in envios_por_destinatario]
    destinatarios_qtd = [item.get('total', 0) for item in envios_por_destinatario]
    context['destinatarios'] = json.dumps(destinatarios)
    context['qtd_destinatarios'] = json.dumps(destinatarios_qtd)

    # ===== Gráfico 4: Gastos por unidade destinatária =====
    gastos_por_destinatario = (
        rateios
        .filter(etiqueta__destinatario__isnull=False)
        .values('etiqueta__destinatario__shopping')
        .annotate(total_gasto=Sum('valor_liquido'))
        .order_by('-total_gasto')
    )
    unidades_destino = [item.get('etiqueta__destinatario__shopping') or 'Não informado' for item in gastos_por_destinatario]
    valores_gastos = [float(item.get('total_gasto') or 0) for item in gastos_por_destinatario]
    context['unidades_destino'] = json.dumps(unidades_destino)
    context['valores_gastos'] = json.dumps(valores_gastos)

    # ===== Gráfico 5: Quantidade de envios por usuário =====
    envios_por_usuario = (
        envios
        .values('user__username')
        .annotate(total=Count('etiqueta'))
        .order_by('-total')
    )
    usuarios = [item.get('user__username') or 'Desconhecido' for item in envios_por_usuario]
    qtd_envios_usuario = [item.get('total') for item in envios_por_usuario]
    context['usuarios'] = json.dumps(usuarios)                # <=== faltava
    context['qtd_envios_usuario'] = json.dumps(qtd_envios_usuario)  # <=== faltava

    # ===== Listas para filtros =====
    context['unidades'] = Unidade.objects.all().order_by('shopping')
    context['usuarios_todos'] = Usuario.objects.select_related('user').all().order_by('user__username')

    return render(request, 'SRCs/graficos.html', context)

@login_required
@permission_required('SRCs.acompanhar_envio', raise_exception=True)
def acompanhamento(request):
    usuario = Usuario.objects.filter(user=request.user).first()

    if not usuario:
        messages.error(request, "Usuário sem vínculo com unidade.")
        return redirect('home')
    
    unidade_usuario = usuario.unidade

    envios = Envio.objects.filter(
        Q(remetente=unidade_usuario) |
        (Q(destinatario=unidade_usuario) & Q(usuario_destinatario__in=[usuario, None])) |
        Q(user=request.user)   # <<< aqui era o problema
    ).order_by('-data_solicitacao')
    context = {
        'envios': envios,
        'usuario': usuario
    }

    return render(request, 'SRCs/acompanhamento.html', context)

@login_required
def detalhe_envio(request, etiqueta):
    envio = get_object_or_404(Envio, etiqueta=etiqueta)
    usuario = Usuario.objects.filter(user=request.user).first()

    if not usuario:
        messages.error(request, "Usuário sem vínculo com unidade.")
        return redirect('acompanhamento')
    
    is_remetente = envio.remetente == usuario.unidade
    is_destinatario = envio.destinatario == usuario.unidade and (
    envio.usuario_destinatario == usuario or envio.usuario_destinatario is None
    )

    if request.method == 'POST':
        if is_remetente and envio.status == 'pendente_envio':
            data_postagem = request.POST.get('data_postagem')
            previsao_chegada = request.POST.get('previsao_chegada')

            if not data_postagem:
                messages.error(request, "A data da postagem é obrigatória.")
            else:
                envio.data_postagem = data_postagem
                envio.previsao_chegada = previsao_chegada or None
                envio.status = 'aguardando_recebimento'
                envio.save()
                messages.success(request, "Informações atualizadas com sucesso.")
                return redirect('detalhe_envio', etiqueta=envio.etiqueta)

        elif is_destinatario and envio.status == 'aguardando_recebimento':
            data_chegada = request.POST.get('data_chegada')

            if not data_chegada:
                messages.error(request, "A data da chegada é obrigatória.")
            else:
                envio.data_chegada = data_chegada
                envio.status = 'entregue'
                envio.save()
                messages.success(request, "Recebimento confirmado com sucesso.")
                return redirect('detalhe_envio', etiqueta=envio.etiqueta)


    context = {
        'envio': envio,
        'usuario': usuario,
        'is_remetente': is_remetente,
        'is_destinatario': is_destinatario,
    }

    return render(request, 'SRCs/detalhe_envio.html', context)

@csrf_exempt
def editar_unidade_ajax(request, unidade_id):
    unidade = get_object_or_404(Unidade, id=unidade_id)
    if request.method == "POST":
        unidade.shopping = request.POST.get('shopping')
        unidade.cnpj = request.POST.get('cnpj')
        unidade.empresa = request.POST.get('empresa')
        unidade.centro_custo = request.POST.get('centro_custo')
        unidade.rua = request.POST.get('rua')
        unidade.bairro = request.POST.get('bairro')
        unidade.cep = request.POST.get('cep')
        unidade.numero = request.POST.get('numero')
        unidade.numero_unidade = request.POST.get('numero_unidade')
        unidade.regional = request.POST.get('regional')
        unidade.cidade = request.POST.get('cidade')
        unidade.estado = request.POST.get('estado')
        unidade.save()
        return JsonResponse({'message': 'Unidade editada com sucesso!'})
    return JsonResponse({'message': 'Método não permitido'}, status=405)

def detalhes_unidade(request, unidade_id):
    unidade = get_object_or_404(Unidade, id=unidade_id)
    data = {
        "id": unidade.id,
        "shopping": unidade.shopping,
        "cnpj": unidade.cnpj,
        "empresa": unidade.empresa,
        "centro_custo": unidade.centro_custo,
        "rua": unidade.rua,
        "bairro": unidade.bairro,
        "cep": unidade.cep,
        "numero": unidade.numero,
        "numero_unidade": unidade.numero_unidade,
        "regional": unidade.regional,
        "cidade": unidade.cidade,
        "estado": unidade.estado,

    }
    return JsonResponse(data)

@csrf_exempt
def excluir_unidades_ajax(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
            ids = data.get("ids", [])
            if not ids:
                return JsonResponse({"error": "Nenhuma unidade selecionada."}, status=400)
            Unidade.objects.filter(id__in=ids).update(excluida=True, cnpj=None)
            return JsonResponse({"message": "Unidade(s) excluída(s) com sucesso!"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Método não permitido."}, status=405)

@csrf_exempt
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)

    if request.method != 'POST':
        return JsonResponse({"error": "Método não permitido."}, status=405)

    username = request.POST.get('username')
    email = request.POST.get('email')
    password = request.POST.get('password')
    perfil = request.POST.get('perfil')
    cartao_postagem = request.POST.get('cartao_postagem')
    unidade_id = request.POST.get('unidade')

    # Valida duplicidade
    if Usuario.objects.filter(user__username=username).exclude(id=usuario.id).exists():
        return JsonResponse({"error": "Username já está em uso."}, status=400)
    if Usuario.objects.filter(user__email=email).exclude(id=usuario.id).exists():
        return JsonResponse({"error": "Email já está em uso."}, status=400)

    # Atualiza dados do User
    usuario.user.username = username
    usuario.user.email = email
    if password:
        usuario.user.set_password(password)
    usuario.user.save()

    # Atualiza dados do Usuario
    usuario.perfil = perfil
    usuario.cartao_postagem = cartao_postagem

    # Atualiza unidade vinculada
    if unidade_id:
        try:
            usuario.unidade = Unidade.objects.get(id=unidade_id)
        except Unidade.DoesNotExist:
            return JsonResponse({"error": "Unidade não encontrada."}, status=400)
    else:
        usuario.unidade = None

    usuario.save()
    return JsonResponse({'message': 'Usuário atualizado com sucesso!'})

@csrf_exempt
def excluir_usuario(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            ids = data.get("ids", [])

            if not ids:
                return JsonResponse({"error": "Nenhum usuário selecionado."}, status=400)

            usuarios = Usuario.objects.filter(id__in=ids)

            for usuario in usuarios:
                user = usuario.user
                if user:
                    # Desativa o usuário
                    user.is_active = False

                    # Adiciona sufixo se ainda não existir
                    sufixo = "(excluido)"
                    if not user.username.endswith(sufixo):
                        user.username = f"{user.username}{sufixo}"
                    if not user.email.endswith(sufixo):
                        user.email = f"{user.email}{sufixo}"

                    user.save()

                # Marca como inativo no seu modelo
                usuario.ativo = False
                usuario.save()

            return JsonResponse({"message": "Usuário(s) desativado(s) com sucesso!"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método não permitido."}, status=405)

def listar_usuarios(request):
    usuarios = Usuario.objects.filter(ativo=True).select_related('user', 'unidade')  # Evita N+1 queries
    return render(request, 'SRCs/listar_usuarios.html', {'usuarios': usuarios})

def listar_unidades(request):
    unidades = Unidade.objects.filter(excluida=False) 
    data = [
        {"id": u.id, "nome": u.shopping}
        for u in unidades
    ]
    return JsonResponse(data, safe=False)

def detalhes_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    return JsonResponse({
        'id': usuario.id,
        'username': usuario.user.username,
        'email': usuario.user.email,
        'perfil': usuario.perfil,
        'cartao_postagem': usuario.cartao_postagem,
        'unidade': usuario.unidade.id if usuario.unidade else '',  
        'unidade_nome': usuario.unidade.shopping if usuario.unidade else '',  
    })

def alterar_foto(request):
    if request.method == "POST" and request.FILES.get('foto'):
        usuario = request.user.usuario
        usuario.foto = request.FILES['foto']
        usuario.save()
        messages.success(request, "Foto atualizada com sucesso!")
    return redirect(request.META.get('HTTP_REFERER', '/'))

